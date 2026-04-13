import io, os, json, webbrowser
from datetime import datetime
import requests
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import msal

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ============================================================
#  CONFIGURATION
# ============================================================

# --- Fichier SOURCE (site entreprise — lecture seule) ---
SP_SOURCE_SITE     = "roseblanchetn.sharepoint.com"
SP_SOURCE_PATH     = "/sites/SDAHSESTPA"
SP_SOURCE_BASE_URL = "https://roseblanchetn.sharepoint.com"
FILE_UNIQUE_ID     = "0761FA65-3D84-4B10-B009-8CA5BF050C98"

# --- Destination : ton OneDrive personnel ---
# ID du document cible extrait de l'URL fournie
# sourcedoc={e7159317-9794-43f3-81fd-6e5ad8b6baca}
ONEDRIVE_HOST      = "roseblanchetn-my.sharepoint.com"
ONEDRIVE_USER      = "siwar_chaari_rose-blanche_com"   # remplace _ par . si besoin
# Dossier de destination dans ton OneDrive (sera créé automatiquement s'il n'existe pas)
ONEDRIVE_FOLDER    = "Anomalies_Export"

# --- Auth ---
CLIENT_ID  = "14d82eec-204b-4c2f-b7e8-296a70dab67e"
AUTHORITY  = "https://login.microsoftonline.com/organizations"
SCOPES_SHAREPOINT = [
    "https://roseblanchetn.sharepoint.com/AllSites.Read",
]
SCOPES_GRAPH = [
    "https://graph.microsoft.com/Files.ReadWrite",
    "https://graph.microsoft.com/User.Read",
]

MOIS_FR = {1:"Janvier",2:"Février",3:"Mars",4:"Avril",5:"Mai",6:"Juin",
           7:"Juillet",8:"Août",9:"Septembre",10:"Octobre",11:"Novembre",12:"Décembre"}

# ============================================================
#  CONFIGURATION DES TABLES
# ============================================================

SSSE_NUMERIC = [
    ("Humidité (%)",               "Teneur en Eau",   lambda v: v < 13.5 or v > 14.5, "13.5 < x < 14.5"),
    ("AW",                         "AW",              lambda v: v > 0.7,              "< 0.7"),
    ("Protéine Brut (%) (+/-0,7)", "Protéine Brute",  lambda v: v < 10,               "> 10 %"),
    ("Protéine (%)/MS",            "Protéine/MS",     lambda v: v < 12,               "> 12 %"),
    ("∑ >400µ",                    "G>400µ",          lambda v: v > 10,               "< 10 %"),
    ("∑ 355;250",                  "G 355-250µ",      lambda v: v < 40,               "> 40 %"),
    ("∑ < 200µ",                   "G<200µ",          lambda v: v > 50,               "< 50 %"),
    ("G < 125µ",                   "G<125µ",          lambda v: v > 10,               "< 10 %"),
    ("Gluten Humide",              "Gluten Humide",   lambda v: v < 28,               "> 28 %"),
    ("Gluten Index",               "Gluten Index",    lambda v: v < 65 or v > 90,    "65 < x < 90"),
    ("Gluten Sec",                 "Gluten Sec",      lambda v: v < 10,               "> 10 %"),
    ("Col. b",                     "Couleur b",       lambda v: v < 18,               "> 18"),
    ("Piqûre Noir",                "Piqûre Noir",     lambda v: v > 10,               "< 10"),
    ("Piqûre Brun",                "Piqûre Brun",     lambda v: v > 100,              "< 100"),
    ("Cendres (%) (+/- 0,02)",     "Cendres",         lambda v: v > 1,                "< 1 %"),
    ("T Chute",                    "Temps de Chute",  lambda v: v < 250,              "> 250"),
]
SSSE_STRING = [
    ("Embalage (Etanchité,visuel...)", "Emballage",  "C"),
    ("C.Poids",                        "Poids",      "C"),
    ("C .Date",                        "Etiquetage", "C"),
]

PS7_NUMERIC = [
    ("Humidité (%)",               "Teneur en Eau",    lambda v: v < 13.5 or v > 14.5, "13.5 < x < 14.5"),
    ("AW",                         "AW",               lambda v: v > 0.7,              "< 0.7"),
    ("Protéine Brut (%) (+/-0,6)", "Protéine Brute",   lambda v: v < 9.6,              "> 9.6 %"),
    ("Prot (%)/MS",                "Protéine/MS",      lambda v: v < 11,               "> 11 %"),
    ("Tps de Chute",               "Temps de Chute",   lambda v: v < 250,              "> 250"),
    ("Amidon End",                 "Amidon End (UCD)", lambda v: v < 18 or v > 25,    "18-25"),
    ("G 200µ",                     "G 200µ",           lambda v: v > 2,                "< 2 %"),
    ("∑ 180;63",                   "∑ 180-63µ",        lambda v: v <= 63,              "> 63 %"),
    ("Gluten Humide",              "Gluten Humide",    lambda v: v < 22,               "> 22 %"),
    ("Col. L",                     "Couleur L",        lambda v: v < 90,               "> 90"),
    ("Cendres (%) (+/- 0,02)",     "Cendres",          lambda v: v > 0.6,              "< 0.6 %"),
    ("Alvéo W",                    "Alvéo W",          lambda v: v < 150,              ">= 150"),
    ("Alvéo P/L",                  "Alvéo P/L",        lambda v: v < 1 or v > 1.8,    "1 <= x <= 1.8"),
    ("Alvéo Ie",                   "Alvéo Ie",         lambda v: v < 45,               "> 45"),
]
PS7_STRING = [
    ("Embalage (Etanchité,visuel...)", "Emballage",  "C"),
    ("C.Poids",                        "Poids",      "C"),
    ("C.Date",                         "Etiquetage", "C"),
]

PS_NUMERIC = [
    ("Humidité (%)",               "Teneur en Eau",    lambda v: v < 13.5 or v > 14.5, "13.5 < x < 14.5"),
    ("AW",                         "AW",               lambda v: v > 0.7,              "< 0.7"),
    ("Protéine Brut (%) (+/-0,6)", "Protéine Brute",   lambda v: v < 9.6,              "> 9.6 %"),
    ("Prot (%)/MS",                "Protéine/MS",      lambda v: v < 11,               "> 11 %"),
    ("Tps de Chute",               "Temps de Chute",   lambda v: v < 250,              "> 250"),
    ("Amidon End",                 "Amidon End (UCD)", lambda v: v < 18 or v > 25,    "18-25"),
    ("G 200µ",                     "G 200µ",           lambda v: v > 2,                "< 2 %"),
    ("∑ 180;63",                   "∑ 180-63µ",        lambda v: v <= 63,              "> 63 %"),
    ("Gluten Humide",              "Gluten Humide",    lambda v: v < 22,               "> 22 %"),
    ("Col. L",                     "Couleur L",        lambda v: v < 89,               "> 89"),
    ("Cendres (%) (+/- 0,02)",     "Cendres",          lambda v: v > 0.7,              "< 0.7 %"),
    ("Alvéo W",                    "Alvéo W",          lambda v: v < 150,              ">= 150"),
    ("Alvéo P/L",                  "Alvéo P/L",        lambda v: v < 1 or v > 1.5,    "1 <= x <= 1.5"),
    ("Alvéo Ie",                   "Alvéo Ie",         lambda v: v < 45,               "> 45"),
]
PS_STRING = [
    ("Emballage (Etanchité,visuel...)", "Emballage",  "C"),
    ("Poids",                           "Poids",      "C"),
    ("C.Date",                          "Etiquetage", "C"),
]

TABLES = [
    {
        "name"      : "SSSE",
        "sheet"     : "Semoule SSSE",
        "output"    : "anomalies_ssse.xlsx",
        "col_date"  : "Date",
        "col_lot"   : "N°lot",
        "col_etape" : "Etape",
        "col_echant": "N° de l'échantillon",
        "col_notif" : "Notif",
        "col_prob"  : None,
        "numeric"   : SSSE_NUMERIC,
        "string"    : SSSE_STRING,
    },
    {
        "name"      : "PS-7",
        "sheet"     : "PS-7",
        "output"    : "anomalies_ps7.xlsx",
        "col_date"  : "Date",
        "col_lot"   : "N°lot",
        "col_etape" : "Etape",
        "col_echant": "N° échantillon",
        "col_notif" : "Résultat",
        "col_prob"  : "Probléme",
        "numeric"   : PS7_NUMERIC,
        "string"    : PS7_STRING,
    },
    {
        "name"      : "PS",
        "sheet"     : "PS",
        "output"    : "anomalies_ps.xlsx",
        "col_date"  : "Date",
        "col_lot"   : "N°lot",
        "col_etape" : "Etape",
        "col_echant": "N° Echantillon",
        "col_notif" : "Résultat",
        "col_prob"  : "Problémes",
        "numeric"   : PS_NUMERIC,
        "string"    : PS_STRING,
    },
]

PARAM_COLORS = {
    "Piqûre Brun"       : "FDECEA",
    "Piqûre Noir"       : "EDE7F6",
    "G>400µ"            : "FFF8E1",
    "G 355-250µ"        : "FFF8E1",
    "G<200µ"            : "FFF8E1",
    "G<125µ"            : "FFF8E1",
    "G 200µ"            : "FFF8E1",
    "∑ 180-63µ"         : "FFF8E1",
    "Couleur b"         : "E3F2FD",
    "Couleur L"         : "E3F2FD",
    "Teneur en Eau"     : "E8F5E9",
    "Gluten Humide"     : "FFF3E0",
    "Gluten Index"      : "FFF3E0",
    "Gluten Sec"        : "FFF3E0",
    "AW"                : "FDECEA",
    "Cendres"           : "F5F5F5",
    "Temps de Chute"    : "F5F5F5",
    "Alvéo W"           : "E8EAF6",
    "Alvéo P/L"         : "E8EAF6",
    "Alvéo Ie"          : "E8EAF6",
    "Protéine Brute"    : "F3E5F5",
    "Protéine/MS"       : "F3E5F5",
    "Amidon End (UCD)"  : "FFF8E1",
    "Emballage"         : "FCE4EC",
    "Poids"             : "FCE4EC",
    "Etiquetage"        : "FCE4EC",
}

# ============================================================
#  ETAPE 1 — Authentification
#  On supprime .token_cache.json si les scopes ont changé
# ============================================================
def _save_cache(cache, pathx):
    if cache.has_state_changed:
        with open(pathx, "w") as f:
            f.write(cache.serialize())

def _get_app(cache_file):
    cache = msal.SerializableTokenCache()
    if os.path.exists(cache_file):
        with open(cache_file) as f:
            cache.deserialize(f.read())
    app = msal.PublicClientApplication(client_id=CLIENT_ID, authority=AUTHORITY, token_cache=cache)
    return app, cache

def _device_flow_login(app, scopes, cache, cache_file, label):
    flow = app.initiate_device_flow(scopes=scopes)
    if "user_code" not in flow:
        raise Exception(f"Erreur flow {label} : {flow}")
    print("\n" + "="*55)
    print(f"  CONNEXION REQUISE ({label})")
    print("="*55)
    print(f"\n  1. Ouvre : https://microsoft.com/devicelogin")
    print(f"  2. Entre le code : {flow['user_code']}")
    print(f"  3. Connecte-toi avec : siwar.chaari@rose-blanche.com")
    print(f"\n  En attente...\n")
    try:
        webbrowser.open("https://microsoft.com/devicelogin")
    except:
        pass
    result = app.acquire_token_by_device_flow(flow)
    if "access_token" not in result:
        raise Exception(f"Connexion échouée ({label}) : {result.get('error_description', result)}")
    print(f"  Connecté ({label}) !")
    _save_cache(cache, cache_file)
    return result["access_token"]

def get_token_sharepoint():
    cache_file = ".token_cache_sp.json"
    app, cache = _get_app(cache_file)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES_SHAREPOINT, account=accounts[0])
        if result and "access_token" in result:
            print("  [SharePoint] Token en cache valide")
            _save_cache(cache, cache_file)
            return result["access_token"]
    return _device_flow_login(app, SCOPES_SHAREPOINT, cache, cache_file, "SharePoint source")

def get_token_graph():
    cache_file = ".token_cache_graph.json"
    app, cache = _get_app(cache_file)
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES_GRAPH, account=accounts[0])
        if result and "access_token" in result:
            print("  [Graph/OneDrive] Token en cache valide")
            _save_cache(cache, cache_file)
            return result["access_token"]
    return _device_flow_login(app, SCOPES_GRAPH, cache, cache_file, "Graph/OneDrive")

def _save_cache(cache, path):
    if cache.has_state_changed:
        with open(path, "w") as f:
            f.write(cache.serialize())

# ============================================================
#  ETAPE 2 — Lecture du fichier SOURCE (site entreprise)
#  Utilise l'URL de téléchargement SharePoint classique.
#  Le fichier original n'est jamais modifié.
# ============================================================
def read_workbook(token) -> bytes:
    """Télécharge le fichier Excel source depuis SharePoint entreprise (lecture seule)."""
    print("  Téléchargement du fichier source (site entreprise)...")
    download_url = (
        f"{SP_SOURCE_BASE_URL}/sites/SDAHSESTPA"
        f"/_layouts/15/download.aspx"
        f"?UniqueId={FILE_UNIQUE_ID}"
    )
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept"       : "*/*",
        "User-Agent"   : "Mozilla/5.0"
    }
    resp = requests.get(download_url, headers=headers, timeout=60, allow_redirects=True)
    if resp.status_code != 200 or b"<!DOCTYPE" in resp.content[:200]:
        raise Exception(f"Erreur {resp.status_code} — supprime .token_cache.json et relance")
    print("  Fichier source téléchargé avec succès (original non modifié).")
    return resp.content

def read_sheet(content: bytes, table_cfg: dict) -> pd.DataFrame:
    try:
        df = pd.read_excel(io.BytesIO(content), sheet_name=table_cfg["sheet"], header=0)
        df.columns = [str(c).strip() for c in df.columns]
        print(f"  [{table_cfg['name']}] {len(df)} lignes lues — feuille '{table_cfg['sheet']}'")
        return df
    except Exception as e:
        try:
            xls = pd.ExcelFile(io.BytesIO(content))
            print(f"  [ERREUR] Feuille '{table_cfg['sheet']}' introuvable!")
            print(f"  [INFO]  Feuilles disponibles : {xls.sheet_names}")
        except:
            pass
        raise

# ============================================================
#  ETAPE 3 — Détection des anomalies
# ============================================================
def _to_float(val):
    if val is None:
        return None
    if isinstance(val, float) and np.isnan(val):
        return None
    try:
        return float(str(val).replace(',', '.').replace(' ', '').strip())
    except (ValueError, TypeError):
        return None

def prepare_data(df: pd.DataFrame, table_cfg: dict):
    col_date   = table_cfg["col_date"]
    col_lot    = table_cfg["col_lot"]
    col_etape  = table_cfg["col_etape"]
    col_echant = table_cfg["col_echant"]
    col_notif  = table_cfg["col_notif"]
    col_prob   = table_cfg["col_prob"]

    df[col_date]  = pd.to_datetime(df[col_date], errors="coerce")
    df[col_etape] = df[col_etape].astype(str).str.strip().str.title()
    df["Année"]   = df[col_date].dt.year.astype("Int64").astype(str)
    df["Mois_num"]= df[col_date].dt.month.astype("Int64")

    rows_anom = []

    for _, r in df.iterrows():
        for col, label, cond, cible in table_cfg["numeric"]:
            v = _to_float(r.get(col))
            if v is not None and cond(v):
                rows_anom.append({
                    "Date"          : r[col_date],
                    "Année"         : r["Année"],
                    "Mois_num"      : r["Mois_num"],
                    "N°lot"         : r.get(col_lot),
                    "N° Echantillon": r.get(col_echant),
                    "Etape"         : r[col_etape],
                    "Parametre"     : label,
                    "Valeur"        : v,
                    "Cible"         : cible,
                    "Notif"         : r.get(col_notif),
                    "Probleme"      : r.get(col_prob) if col_prob else "",
                    "Commentaires"  : r.get("Commentaire", r.get("Commentaires", "")),
                })

        for col, label, target in table_cfg["string"]:
            v = r.get(col)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                if str(v).strip() != target:
                    rows_anom.append({
                        "Date"          : r[col_date],
                        "Année"         : r["Année"],
                        "Mois_num"      : r["Mois_num"],
                        "N°lot"         : r.get(col_lot),
                        "N° Echantillon": r.get(col_echant),
                        "Etape"         : r[col_etape],
                        "Parametre"     : label,
                        "Valeur"        : str(v).strip(),
                        "Cible"         : target,
                        "Notif"         : r.get(col_notif),
                        "Probleme"      : r.get(col_prob) if col_prob else "",
                        "Commentaires"  : r.get("Commentaire", r.get("Commentaires", "")),
                    })

    df_anom = pd.DataFrame(rows_anom)

    print(f"  [{table_cfg['name']}] Total lignes analysées  : {len(df)}")
    print(f"  [{table_cfg['name']}] Total anomalies réelles : {len(df_anom)}")
    if len(df_anom) > 0:
        print(f"  Répartition par paramètre :")
        for param, cnt in df_anom["Parametre"].value_counts().items():
            print(f"    {param:<28} {cnt}")

    return df, df_anom

# ============================================================
#  ETAPE 4 — Génération Excel en mémoire (4 feuilles)
#  Retourne des bytes — aucun fichier local créé.
# ============================================================
def generate_excel_bytes(df_all: pd.DataFrame, df_anom: pd.DataFrame, table_cfg: dict) -> bytes:
    has_prob = table_cfg["col_prob"] is not None

    wb = Workbook()

    HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11, name="Arial")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BORDER = Border(
        left  =Side(style="thin", color="D0D0D0"),
        right =Side(style="thin", color="D0D0D0"),
        top   =Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
            c.alignment = HEADER_ALIGN; c.border = BORDER

    def style_row(ws, row, cols, fill_hex=None):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            if fill_hex:
                c.fill = PatternFill("solid", fgColor=fill_hex)
            c.border = BORDER
            c.alignment = Alignment(vertical="center")

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Feuille 1 : Anomalies_Detail ----
    ws1 = wb.active
    ws1.title = "Anomalies_Detail"
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 35

    h1 = ["Date", "Année", "Mois", "Semaine", "N° Lot", "N° Échantillon",
          "Étape", "Paramètre", "Valeur mesurée", "Cible", "Résultat/Notif"]
    if has_prob:
        h1.append("Problème déclaré")
    h1.append("Commentaires")

    ws1.append(h1)
    style_header(ws1, 1, len(h1))

    for i, (_, r) in enumerate(df_anom.iterrows(), 2):
        d = r["Date"] if pd.notna(r["Date"]) else None
        row_data = [
            d,
            str(r["Année"])      if pd.notna(r.get("Année"))          else "",
            int(r["Mois_num"])   if pd.notna(r.get("Mois_num"))       else "",
            d.isocalendar()[1]   if d else "",
            str(r["N°lot"])          if pd.notna(r.get("N°lot"))          else "",
            str(r["N° Echantillon"]) if pd.notna(r.get("N° Echantillon")) else "",
            str(r["Etape"]),
            str(r["Parametre"]),
            r["Valeur"],
            str(r["Cible"]),
            str(r["Notif"]) if pd.notna(r.get("Notif")) else "",
        ]
        if has_prob:
            row_data.append(str(r["Probleme"]) if pd.notna(r.get("Probleme")) else "")
        row_data.append(str(r["Commentaires"]) if pd.notna(r.get("Commentaires")) else "")

        ws1.append(row_data)
        if d:
            ws1.cell(row=i, column=1).number_format = "DD/MM/YYYY"
        fill = PARAM_COLORS.get(str(r["Parametre"]), "FFFFFF")
        style_row(ws1, i, len(h1), fill_hex=fill if i % 2 == 0 else None)

    widths1 = [13, 8, 7, 9, 10, 16, 16, 22, 15, 16, 14]
    if has_prob:
        widths1.append(22)
    widths1.append(30)
    set_widths(ws1, widths1)

    # ---- Feuille 2 : Resume_Mensuel ----
    ws2 = wb.create_sheet("Resume_Mensuel")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 35

    df2 = df_anom.copy()
    df2["Mois_n"] = df2["Date"].dt.month
    df2["An"]     = df2["Date"].dt.year

    def count_notif(x):
        return sum(
            1 for v in x
            if pd.notna(v) and str(v).strip().lower() in ("oui", "libération", "liberation")
        )

    grp2 = df2.groupby(["An", "Mois_n"]).agg(
        Nb       =("Parametre", "count"),
        Notifiees=("Notif",     count_notif),
        Types    =("Parametre", "nunique")
    ).reset_index().sort_values(["An", "Mois_n"])

    h2 = ["Année", "Mois Num", "Mois", "Nb Anomalies", "Nb Résolues", "Types Distincts", "Taux Résolution (%)"]
    ws2.append(h2)
    style_header(ws2, 1, len(h2))

    for i, (_, r) in enumerate(grp2.iterrows(), 2):
        taux = round(r["Notifiees"] / r["Nb"] * 100, 1) if r["Nb"] > 0 else 0
        ws2.append([int(r["An"]), int(r["Mois_n"]), MOIS_FR.get(int(r["Mois_n"]), ""),
                    int(r["Nb"]), int(r["Notifiees"]), int(r["Types"]), taux])
        style_row(ws2, i, len(h2), fill_hex="F0F4FA" if i % 2 == 0 else None)
    set_widths(ws2, [10, 10, 14, 16, 16, 16, 22])

    if grp2.shape[0] > 0:
        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = f"Anomalies par mois — {table_cfg['name']}"
        chart2.y_axis.title = "Nb Anomalies"
        chart2.add_data(Reference(ws2, min_col=4, min_row=1, max_row=grp2.shape[0]+1), titles_from_data=True)
        chart2.set_categories(Reference(ws2, min_col=3, min_row=2, max_row=grp2.shape[0]+1))
        chart2.width = 20; chart2.height = 12
        ws2.add_chart(chart2, "I2")

    # ---- Feuille 3 : Resume_Parametre ----
    ws3 = wb.create_sheet("Resume_Parametre")
    ws3.row_dimensions[1].height = 35
    total = len(df_anom)

    grp3 = df_anom.groupby("Parametre").agg(
        Nb       =("Parametre", "count"),
        Notifiees=("Notif",     count_notif),
        Etapes   =("Etape",     "nunique")
    ).reset_index().sort_values("Nb", ascending=False)

    cible_map = {label: cible for (_, label, _, cible) in table_cfg["numeric"]}
    cible_map.update({label: target for (_, label, target) in table_cfg["string"]})

    h3 = ["Paramètre", "Cible", "Nb Anomalies", "% du Total", "Nb Résolues", "Taux Résolution (%)", "Étapes Concernées"]
    ws3.append(h3)
    style_header(ws3, 1, len(h3))

    for i, (_, r) in enumerate(grp3.iterrows(), 2):
        pct  = round(r["Nb"] / total * 100, 1) if total > 0 else 0
        taux = round(r["Notifiees"] / r["Nb"] * 100, 1) if r["Nb"] > 0 else 0
        ws3.append([
            str(r["Parametre"]),
            cible_map.get(str(r["Parametre"]), ""),
            int(r["Nb"]), pct,
            int(r["Notifiees"]), taux,
            int(r["Etapes"])
        ])
        fill = PARAM_COLORS.get(str(r["Parametre"]), "F0F4FA")
        style_row(ws3, i, len(h3), fill_hex=fill)
    set_widths(ws3, [25, 16, 14, 12, 14, 22, 18])

    if grp3.shape[0] > 0:
        pie3 = PieChart()
        pie3.title = f"Répartition par paramètre — {table_cfg['name']}"
        pie3.add_data(Reference(ws3, min_col=3, min_row=1, max_row=grp3.shape[0]+1), titles_from_data=True)
        pie3.set_categories(Reference(ws3, min_col=1, min_row=2, max_row=grp3.shape[0]+1))
        pie3.width = 18; pie3.height = 12
        ws3.add_chart(pie3, "I2")

    # ---- Feuille 4 : Resume_Etape ----
    ws4 = wb.create_sheet("Resume_Etape")
    ws4.row_dimensions[1].height = 35

    grp4 = df_anom.groupby("Etape").agg(
        Nb       =("Parametre", "count"),
        Notifiees=("Notif",     count_notif),
        Types    =("Parametre", "nunique")
    ).reset_index().sort_values("Nb", ascending=False)

    h4 = ["Étape", "Nb Anomalies", "% du Total", "Nb Résolues", "Taux Résolution (%)", "Types Distincts"]
    ws4.append(h4)
    style_header(ws4, 1, len(h4))

    for i, (_, r) in enumerate(grp4.iterrows(), 2):
        pct  = round(r["Nb"] / total * 100, 1) if total > 0 else 0
        taux = round(r["Notifiees"] / r["Nb"] * 100, 1) if r["Nb"] > 0 else 0
        ws4.append([str(r["Etape"]), int(r["Nb"]), pct, int(r["Notifiees"]), taux, int(r["Types"])])
        style_row(ws4, i, len(h4), fill_hex="F0F4FA" if i % 2 == 0 else None)
    set_widths(ws4, [18, 14, 12, 14, 22, 16])

    if grp4.shape[0] > 0:
        chart4 = BarChart()
        chart4.type = "bar"
        chart4.title = f"Anomalies par étape — {table_cfg['name']}"
        chart4.add_data(Reference(ws4, min_col=2, min_row=1, max_row=grp4.shape[0]+1), titles_from_data=True)
        chart4.set_categories(Reference(ws4, min_col=1, min_row=2, max_row=grp4.shape[0]+1))
        chart4.width = 18; chart4.height = 12
        ws4.add_chart(chart4, "H2")

    # --- Sauvegarde en mémoire (pas sur disque) ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_bytes = buffer.read()
    print(f"  [{table_cfg['name']}] Excel généré en mémoire ({len(file_bytes)//1024} KB)")
    return file_bytes

# ============================================================
#  ETAPE 5 — Upload vers ton OneDrive personnel (Graph API)
#  Le fichier est placé dans /Anomalies_Export/<filename>
#  dans TON OneDrive personnel, séparé du site entreprise.
# ============================================================
def get_my_drive_id(token: str) -> str:
    """Récupère l'ID du OneDrive de l'utilisateur connecté."""
    resp = requests.get(
        "https://graph.microsoft.com/v1.0/me/drive",
        headers={"Authorization": f"Bearer {token}"}
    )
    if resp.status_code != 200:
        raise Exception(f"Impossible de récupérer le OneDrive : {resp.status_code} — {resp.text}")
    drive = resp.json()
    print(f"  OneDrive personnel : {drive.get('name', '?')} ({drive['id']})")
    return drive["id"]

def ensure_folder(token: str, drive_id: str, folder_name: str) -> str:
    """Crée le dossier de destination s'il n'existe pas. Retourne l'ID du dossier."""
    # Vérifie si le dossier existe déjà dans la racine
    url = f"https://graph.microsoft.com/v1.0/me/drive/root/children"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    resp = requests.get(url, headers=headers)
    items = resp.json().get("value", [])
    for item in items:
        if item.get("name") == folder_name and "folder" in item:
            print(f"  Dossier '{folder_name}' trouvé (id: {item['id']})")
            return item["id"]

    # Crée le dossier
    payload = {
        "name": folder_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "rename"
    }
    resp = requests.post(url, headers=headers, json=payload)
    if resp.status_code not in (200, 201):
        raise Exception(f"Erreur création dossier : {resp.status_code} — {resp.text}")
    folder_id = resp.json()["id"]
    print(f"  Dossier '{folder_name}' créé (id: {folder_id})")
    return folder_id

def upload_to_my_onedrive(token: str, file_bytes: bytes, filename: str) -> str:
    """
    Upload un fichier Excel vers ton OneDrive personnel (roseblanchetn-my.sharepoint.com).
    Utilise l'API Graph /me/drive — complètement séparé du site entreprise.
    Retourne l'URL web du fichier uploadé.
    """
    print(f"  Upload de '{filename}' vers OneDrive personnel...")

    # Pour les fichiers > 4 MB on utilise une upload session, sinon PUT simple
    size_mb = len(file_bytes) / (1024 * 1024)

    headers_base = {"Authorization": f"Bearer {token}"}

    if size_mb < 4:
        # --- Upload simple (PUT) ---
        url = (
            f"https://graph.microsoft.com/v1.0/me/drive/root:/"
            f"{ONEDRIVE_FOLDER}/{filename}:/content"
        )
        headers = {
            **headers_base,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        resp = requests.put(url, headers=headers, data=file_bytes, timeout=60)

        if resp.status_code in (200, 201):
            item    = resp.json()
            web_url = item.get("webUrl", "")
            print(f"  ✅ '{filename}' uploadé sur OneDrive personnel")
            print(f"     URL : {web_url}")
            return web_url
        else:
            raise Exception(f"Erreur upload '{filename}' : {resp.status_code} — {resp.text}")

    else:
        # --- Upload session (pour fichiers > 4 MB) ---
        session_url = (
            f"https://graph.microsoft.com/v1.0/me/drive/root:/"
            f"{ONEDRIVE_FOLDER}/{filename}:/createUploadSession"
        )
        session_payload = {
            "item": {
                "@microsoft.graph.conflictBehavior": "replace",
                "name": filename
            }
        }
        sess_resp = requests.post(
            session_url,
            headers={**headers_base, "Content-Type": "application/json"},
            json=session_payload, timeout=30
        )
        if sess_resp.status_code != 200:
            raise Exception(f"Erreur création session upload : {sess_resp.status_code} — {sess_resp.text}")

        upload_url = sess_resp.json()["uploadUrl"]
        chunk_size = 3 * 1024 * 1024  # 3 MB par chunk
        total      = len(file_bytes)
        start      = 0
        last_resp  = None

        while start < total:
            end   = min(start + chunk_size - 1, total - 1)
            chunk = file_bytes[start:end + 1]
            chunk_headers = {
                "Content-Length": str(len(chunk)),
                "Content-Range" : f"bytes {start}-{end}/{total}",
                "Content-Type"  : "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
            r = requests.put(upload_url, headers=chunk_headers, data=chunk, timeout=60)
            last_resp = r
            start = end + 1
            print(f"     Envoyé {start}/{total} octets...")

        if last_resp and last_resp.status_code in (200, 201):
            item    = last_resp.json()
            web_url = item.get("webUrl", "")
            print(f"  ✅ '{filename}' uploadé sur OneDrive personnel")
            print(f"     URL : {web_url}")
            return web_url
        else:
            status = last_resp.status_code if last_resp else "?"
            raise Exception(f"Erreur upload session '{filename}' : {status}")

# ============================================================
#  MAIN
# ============================================================
if __name__ == "__main__":
    print("\n" + "="*60)
    print("  Export Anomalies SSSE / PS-7 / PS → OneDrive Personnel")
    print("="*60 + "\n")
    print("  ℹ️  Les fichiers originaux sur le site entreprise")
    print("      ne seront PAS modifiés (lecture seule).\n")
    print("  ℹ️  Les fichiers générés seront déposés dans :")
    print(f"      OneDrive personnel → /{ONEDRIVE_FOLDER}/\n")

    # ---- Auth ----
    print("[1/4] Authentification Microsoft...")
    print("      Si tu as déjà un .token_cache.json avec les anciens")
    print("      scopes, supprime-le pour forcer une reconnexion.\n")
    token_sp = get_token_sharepoint()
    token_graph = get_token_graph()

    # ---- Téléchargement source ----
    print("\n[2/4] Téléchargement fichier Excel source (site entreprise)...")
    raw_content = read_workbook(token_sp)

    # ---- Traitement des 3 tables ----
    print("\n[3/4] Détection des anomalies — 3 tables...")
    results = []   # [(filename, file_bytes)]
    errors  = []

    for table_cfg in TABLES:
        try:
            print(f"\n  --- {table_cfg['name']} ---")
            df_raw          = read_sheet(raw_content, table_cfg)
            df_all, df_anom = prepare_data(df_raw, table_cfg)
            file_bytes      = generate_excel_bytes(df_all, df_anom, table_cfg)
            results.append((table_cfg["output"], file_bytes, len(df_anom)))
        except Exception as e:
            err = f"ERREUR [{table_cfg['name']}]: {e}"
            print(f"\n  ❌ {err}")
            errors.append(err)
            import traceback; traceback.print_exc()

    # ---- Upload vers OneDrive personnel ----
    print("\n[4/4] Upload vers OneDrive personnel...")
    uploaded_urls = []

    for filename, file_bytes, nb_anom in results:
        try:
            url = upload_to_my_onedrive(token_graph, file_bytes, filename)
            uploaded_urls.append((filename, url, nb_anom))
        except Exception as e:
            err = f"ERREUR upload [{filename}]: {e}"
            print(f"\n  ❌ {err}")
            errors.append(err)
            import traceback; traceback.print_exc()

    # ---- Résumé ----
    print("\n" + "="*60)
    print("  TERMINÉ")
    print("="*60)

    if errors:
        print("\n  ⚠️  Erreurs rencontrées :")
        for e in errors:
            print(f"    ❌ {e}")

    if uploaded_urls:
        print("\n  ✅ Fichiers déposés sur ton OneDrive personnel :")
        for fname, url, nb in uploaded_urls:
            print(f"\n    📄 {fname}  ({nb} anomalies)")
            print(f"       {url}")

    print(f"\n  Dossier OneDrive : /{ONEDRIVE_FOLDER}/")
    print("\n  Prochaine étape Power BI :")
    print("  → Dans Power BI Desktop : Obtenir données → SharePoint Online")
    print(f"  → URL du site : https://{ONEDRIVE_HOST}/personal/{ONEDRIVE_USER}")
    print("  → Navigue jusqu'au dossier Anomalies_Export")
    print("  → Publie sur Power BI Service et active le rafraîchissement planifié")
    print()